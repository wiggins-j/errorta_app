"""XLSX extraction via openpyxl. One chunk per sheet."""
from __future__ import annotations

from pathlib import Path

from . import Chunk, ExtractError


def extract(path: Path) -> list[Chunk]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:  # pragma: no cover
        raise ExtractError(f"openpyxl not available: {e}") from e

    try:
        wb = load_workbook(str(path), data_only=True, read_only=True)
    except Exception as e:
        raise ExtractError(f"could not open XLSX: {e}") from e

    chunks: list[Chunk] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines = [f"# Sheet: {sheet_name}"]
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            if any(c.strip() for c in cells):
                lines.append("\t".join(cells))
        text = "\n".join(lines).strip()
        if not text:
            continue
        chunks.append(
            {"text": text, "meta": {"sheet": sheet_name, "source_type": "xlsx"}}
        )
    wb.close()
    if not chunks:
        raise ExtractError("XLSX has no non-empty sheets")
    return chunks
